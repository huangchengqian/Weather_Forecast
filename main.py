import json
import os
from openpyxl.workbook import Workbook
from Weather_Forecast import sendEmail, weather_Fore, write_2_Excel

if __name__ == '__main__':
	with open("city.json", 'r') as load_f:
		c_json = json.load(load_f)
	# print(c_json)

	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = "head_city"
	# 创建表头
	worksheet.cell(column=1, row=1, value="城  市")
	worksheet.cell(column=2, row=1, value="日  期")
	worksheet.cell(column=3, row=1, value="天  气")
	worksheet.cell(column=4, row=1, value="平均气温(摄氏度)")
	rows = 2

	weather_info = ""

	for lists in c_json:
		headCity = lists['id']
		if (headCity[5:9] == "0101") | (lists['provinceEn'] == lists['cityEn']):
			# print(lists['cityZh'])
			te = weather_Fore.getInfo(headCity)
			w_json = json.loads(te)
			# 窗口输出
			write_2_Excel.write_and_save(w_json)
			# 写入Excel
			write_2_Excel.write2Excel(worksheet, rows, w_json)
			rows += 1
			weather_info += ("城市:" + w_json['city'] + "  天气:" + w_json["wea"] + "  最高气温:" + w_json['tem1'] + "  最低气温:" + w_json['tem2'] + '\n')

	workbook.save(filename="weather.xlsx")
	# print(weather_info)

	if input("是否通过邮件发送(y/n)?") == 'y':
		sendEmail.send_mail(weather_info)