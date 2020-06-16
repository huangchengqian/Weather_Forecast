from openpyxl.workbook import Workbook


def write2Excel(worksheet, rows, w_json):
	ave_tem = (int(w_json['tem1']) + int(w_json['tem2'])) / 2
	worksheet.cell(column=1, row=rows, value=w_json['city'])
	worksheet.cell(column=2, row=rows, value=w_json["date"])
	worksheet.cell(column=3, row=rows, value=w_json["wea"])
	worksheet.cell(column=4, row=rows, value=ave_tem)


def write_and_save(w_json):
	ave_tem = (int(w_json['tem1']) + int(w_json['tem2'])) / 2
	print("城  市: {0:>4}\t|\t日  期: {1}\t|\t天  气: {2:>2}\t|\t实时气温: {3}\t|\t平均气温: {4:>4}\t|".format(w_json['city'], w_json["date"], w_json["wea"], w_json['tem'], ave_tem))
	print("*********************************************************************************************")